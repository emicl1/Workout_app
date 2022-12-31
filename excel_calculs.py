from openpyxl import load_workbook


def change_value(lift, value):
    workbook = load_workbook(filename="Candito 6 Week Program.xlsx")
    if lift == "bench":
        sheet = workbook.active
        sheet["B12"] = value
    if lift == "squat":
        sheet = workbook.active
        sheet["B13"] = value
    if lift == "deadlift":
        sheet = workbook.active
        sheet["B14"] = value
    else:
        workbook.save(filename="Candito 6 Week Program.xlsx")
        return "Default values are in"
    workbook.save(filename="Candito 6 Week Program.xlsx")


def return_accessory():
    return [{"horizontal_pull": ("Dumbbell Row", "Barbell Row", "Machine Row")},
            {"vertical_pull": ("Weighted Pull-up", "Lat Pulldown", "Weighted Chin-up")},
            {"Shoulder Exercise": ("Standing Dumbbell OHP", "Seated Dumbbell OHP", "Military Press", "Lateral Raise")}]


def change_accessory(typ_of_lift, lift):
    workbook = load_workbook(filename="Candito 6 Week Program.xlsx")
    if typ_of_lift == "horizontal_pull":
        sheet = workbook.active
        sheet["B17"] = lift
    if typ_of_lift == "shoulder_exercise":
        sheet = workbook.active
        sheet["B18"] = lift
    if typ_of_lift == "vertical_pull":
        sheet = workbook.active
        sheet["B19"] = lift
    else:
        workbook.save(filename="Candito 6 Week Program.xlsx")
        return "Default values are in"
    workbook.save(filename="Candito 6 Week Program.xlsx")


def change_data(day, month, year):
    workbook = load_workbook(filename="Candito 6 Week Program.xlsx")
    sheet = workbook.active
    date_format = f"{month}/{day}/{year}"
    sheet["B6"] = date_format
    workbook.save(filename="Candito 6 Week Program.xlsx")


def get_value(lift):
    workbook = load_workbook(filename="Candito 6 Week Program.xlsx")
    if lift == "bench":
        sheet = workbook.active
        return sheet["B12"].value
    if lift == "squat":
        sheet = workbook.active
        return sheet["B13"].value
    if lift == "deadlift":
        sheet = workbook.active
        return sheet["B14"].value
    else:
        return "lift not found"


def traverse_a_sheet(name_of_sheet):
    workbook = load_workbook(filename="Candito 6 Week Program.xlsx", data_only=True)
    sheet = workbook[name_of_sheet]
    for row in sheet.iter_rows(min_row=1, max_col=10, values_only=True):
        for cell in row:
            if cell is not None:
                print(cell)
